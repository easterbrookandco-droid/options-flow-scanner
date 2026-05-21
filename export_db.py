# -*- coding: utf-8 -*-
"""
export_db.py

Exports all key tables from signals.db to an Excel workbook
for manual review and validation.

Sheets included:
  1. paper_trades      — all trades with full entry/exit details
  2. open_positions    — currently open and stop-triggered positions
  3. closed_summary    — closed trades with P&L by exit reason
  4. signals_today     — today's scanner signals
  5. market_closes     — stored previous close prices

Usage:
    python export_db.py
"""

import sqlite3
import os
import pytz
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

MARKET_TIMEZONE = "US/Eastern"

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl --break-system-packages")
    exit(1)


def get_db():
    from journal import DB_PATH
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def style_header_row(ws, row_num, color="1F4E79"):
    """Apply header styling to a row."""
    for cell in ws[row_num]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws):
    """Auto-size all columns based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def color_pnl_cells(ws, pnl_col_idx, start_row=2):
    """Color P&L cells green/red based on value."""
    green = PatternFill("solid", fgColor="C6EFCE")
    red   = PatternFill("solid", fgColor="FFC7CE")

    for row in ws.iter_rows(min_row=start_row, min_col=pnl_col_idx,
                             max_col=pnl_col_idx):
        for cell in row:
            if cell.value is not None:
                try:
                    val = float(cell.value)
                    if val > 0:
                        cell.fill = green
                    elif val < 0:
                        cell.fill = red
                except (ValueError, TypeError):
                    pass


def export_paper_trades(wb, conn):
    """Sheet 1 — All paper trades."""
    ws = wb.create_sheet("All Trades")

    cursor = conn.cursor()
    cursor.execute("""
        SELECT
            id,
            signal_contract,
            entry_date,
            exit_date,
            status,
            exit_reason,
            entry_price,
            exit_price,
            contracts,
            total_cost,
            pnl,
            pnl_pct,
            hold_days,
            score_at_entry,
            dte_at_entry,
            iv_at_entry,
            delta_at_entry,
            verdict_at_entry,
            market_bias_at_entry,
            ticker_bias_at_entry,
            target_price,
            stop_price,
            thesis,
            notes
        FROM paper_trades
        ORDER BY id DESC
    """)

    headers = [
        "ID", "Contract", "Entry Date", "Exit Date", "Status",
        "Exit Reason", "Entry Price", "Exit Price", "Contracts",
        "Total Cost", "P&L", "P&L %", "Hold Days",
        "Score", "DTE", "IV", "Delta", "Verdict",
        "Market Bias", "Ticker Bias", "Target", "Stop",
        "Thesis", "Notes"
    ]

    ws.append(headers)
    style_header_row(ws, 1)

    rows = cursor.fetchall()
    for row in rows:
        ws.append(list(row))

    # Color P&L column (index 11 = column K)
    color_pnl_cells(ws, 11)

    auto_width(ws)
    ws.freeze_panes = "A2"
    return len(rows)


def export_open_positions(wb, conn):
    """Sheet 2 — Open and stop-triggered positions with last snapshot."""
    ws = wb.create_sheet("Open Positions")

    cursor = conn.cursor()
    cursor.execute("""
        WITH last_snapshots AS (
            SELECT trade_id, MAX(id) as max_id
            FROM position_snapshots
            GROUP BY trade_id
        )
        SELECT
            pt.id,
            pt.signal_contract,
            pt.status,
            pt.entry_date,
            pt.entry_price,
            pt.contracts,
            pt.total_cost,
            pt.target_price,
            pt.score_at_entry,
            pt.dte_at_entry,
            ps.current_price as last_price,
            ps.pnl           as unrealized_pnl,
            ps.pnl_pct       as unrealized_pnl_pct,
            ps.current_dte,
            ps.dynamic_stop,
            ps.snapshot_time as last_snapshot,
            ps.spy_chg_pct,
            ps.qqq_chg_pct,
            ps.vix_price,
            pt.thesis
        FROM paper_trades pt
        JOIN last_snapshots ls ON ls.trade_id = pt.id
        JOIN position_snapshots ps ON ps.id = ls.max_id
        WHERE pt.status IN ('OPEN', 'STOP_TRIGGERED')
        ORDER BY ps.pnl DESC
    """)

    headers = [
        "ID", "Contract", "Status", "Entry Date", "Entry Price",
        "Contracts", "Total Cost", "Target", "Score", "DTE at Entry",
        "Last Price", "Unrealized P&L", "Unrealized P&L %",
        "Current DTE", "Dynamic Stop", "Last Snapshot",
        "SPY Chg%", "QQQ Chg%", "VIX", "Thesis"
    ]

    ws.append(headers)
    style_header_row(ws, 1, color="375623")

    rows = cursor.fetchall()
    for row in rows:
        ws.append(list(row))

    # Color unrealized P&L column (index 12 = column L)
    color_pnl_cells(ws, 12)

    auto_width(ws)
    ws.freeze_panes = "A2"
    return len(rows)


def export_closed_summary(wb, conn):
    """Sheet 3 — Closed trade summary by various dimensions."""
    ws = wb.create_sheet("Closed Summary")

    cursor = conn.cursor()

    row = 1

    # ── Overall summary ────────────────────────────────────────────────
    ws.cell(row, 1, "OVERALL SUMMARY").font = Font(bold=True, size=12)
    row += 1

    cursor.execute("""
        SELECT
            COUNT(*)                                        as total,
            SUM(CASE WHEN pnl > 0 THEN 1 ELSE 0 END)      as wins,
            SUM(CASE WHEN pnl <= 0 THEN 1 ELSE 0 END)     as losses,
            ROUND(SUM(pnl), 2)                             as total_pnl,
            ROUND(AVG(pnl), 2)                             as avg_pnl,
            ROUND(AVG(pnl_pct), 1)                        as avg_pnl_pct,
            MAX(pnl)                                       as best_trade,
            MIN(pnl)                                       as worst_trade
        FROM paper_trades WHERE status = 'CLOSED'
    """)
    summary = cursor.fetchone()
    total   = summary[0] or 0
    wins    = summary[1] or 0
    wr      = round((wins / total * 100), 1) if total else 0

    ws.append(["Metric", "Value"])
    style_header_row(ws, row)
    row += 1

    for label, val in [
        ("Total Closed", total),
        ("Wins", wins),
        ("Losses", summary[2]),
        ("Win Rate", f"{wr}%"),
        ("Total P&L", summary[3]),
        ("Avg P&L", summary[4]),
        ("Avg P&L %", f"{summary[5]}%"),
        ("Best Trade", summary[6]),
        ("Worst Trade", summary[7]),
    ]:
        ws.append([label, val])
        row += 1

    row += 1

    # ── By exit reason ─────────────────────────────────────────────────
    ws.cell(row, 1, "BY EXIT REASON").font = Font(bold=True, size=12)
    row += 1

    headers = ["Exit Reason", "Count", "Wins", "Losses", "Win %",
               "Total P&L", "Avg P&L", "Avg P&L %"]
    ws.append(headers)
    style_header_row(ws, row)
    row += 1

    cursor.execute("""
        SELECT exit_reason,
               COUNT(*) as count,
               SUM(CASE WHEN pnl > 0 THEN 1 ELSE 0 END) as wins,
               SUM(CASE WHEN pnl <= 0 THEN 1 ELSE 0 END) as losses,
               ROUND(SUM(pnl), 2) as total_pnl,
               ROUND(AVG(pnl), 2) as avg_pnl,
               ROUND(AVG(pnl_pct), 1) as avg_pct
        FROM paper_trades WHERE status = 'CLOSED'
        GROUP BY exit_reason ORDER BY total_pnl DESC
    """)
    for r in cursor.fetchall():
        cnt = r[1] or 0
        w   = r[2] or 0
        wr2 = round((w / cnt * 100), 1) if cnt else 0
        ws.append([r[0], cnt, w, r[3], f"{wr2}%", r[4], r[5], f"{r[6]}%"])
        row += 1

    row += 1

    # ── By ticker ──────────────────────────────────────────────────────
    ws.cell(row, 1, "BY TICKER").font = Font(bold=True, size=12)
    row += 1

    headers = ["Ticker", "Count", "Wins", "Losses", "Win %",
               "Total P&L", "Avg P&L %"]
    ws.append(headers)
    style_header_row(ws, row)
    row += 1

    cursor.execute("""
        SELECT
            CASE
                WHEN signal_contract LIKE 'SPY%' THEN 'SPY'
                WHEN signal_contract LIKE 'QQQ%' THEN 'QQQ'
                WHEN signal_contract LIKE 'TSLA%' THEN 'TSLA'
                WHEN signal_contract LIKE 'NVDA%' THEN 'NVDA'
                WHEN signal_contract LIKE 'NFLX%' THEN 'NFLX'
                WHEN signal_contract LIKE 'AMD%' THEN 'AMD'
                WHEN signal_contract LIKE 'META%' THEN 'META'
                WHEN signal_contract LIKE 'AAPL%' THEN 'AAPL'
                WHEN signal_contract LIKE 'MSFT%' THEN 'MSFT'
                WHEN signal_contract LIKE 'AMZN%' THEN 'AMZN'
                WHEN signal_contract LIKE 'GOOGL%' THEN 'GOOGL'
                WHEN signal_contract LIKE 'IWM%' THEN 'IWM'
                WHEN signal_contract LIKE 'CRM%' THEN 'CRM'
                WHEN signal_contract LIKE 'JPM%' THEN 'JPM'
                WHEN signal_contract LIKE 'GS%' THEN 'GS'
                WHEN signal_contract LIKE 'BAC%' THEN 'BAC'
                WHEN signal_contract LIKE 'UBER%' THEN 'UBER'
                WHEN signal_contract LIKE 'XLF%' THEN 'XLF'
                WHEN signal_contract LIKE 'XLE%' THEN 'XLE'
                ELSE 'OTHER'
            END as ticker,
            COUNT(*) as count,
            SUM(CASE WHEN pnl > 0 THEN 1 ELSE 0 END) as wins,
            SUM(CASE WHEN pnl <= 0 THEN 1 ELSE 0 END) as losses,
            ROUND(SUM(pnl), 2) as total_pnl,
            ROUND(AVG(pnl_pct), 1) as avg_pct
        FROM paper_trades WHERE status = 'CLOSED'
        GROUP BY ticker ORDER BY total_pnl DESC
    """)
    for r in cursor.fetchall():
        cnt = r[1] or 0
        w   = r[2] or 0
        wr2 = round((w / cnt * 100), 1) if cnt else 0
        ws.append([r[0], cnt, w, r[3], f"{wr2}%", r[4], f"{r[5]}%"])
        row += 1

    auto_width(ws)
    return row


def export_signals_today(wb, conn):
    """Sheet 4 — Today's scanner signals."""
    ws = wb.create_sheet("Today's Signals")

    eastern = pytz.timezone(MARKET_TIMEZONE)
    today   = datetime.now(eastern).strftime("%Y-%m-%d")

    cursor = conn.cursor()
    cursor.execute("""
        SELECT
            id, scan_time, ticker, contract, contract_type,
            strike, expiration, signal_tier, composite_score,
            premium, bid, ask, volume, open_interest,
            vol_oi_ratio, share_price, outcome
        FROM signals
        WHERE scan_time LIKE ?
        ORDER BY composite_score DESC
    """, (f"{today}%",))

    headers = [
        "ID", "Scan Time", "Ticker", "Contract", "Type",
        "Strike", "Expiry", "Tier", "Score", "Premium",
        "Bid", "Ask", "Volume", "OI", "Vol/OI",
        "Share Price", "Outcome"
    ]

    ws.append(headers)
    style_header_row(ws, 1, color="6B2737")

    rows = cursor.fetchall()
    for row in rows:
        ws.append(list(row))

    auto_width(ws)
    ws.freeze_panes = "A2"
    return len(rows)


def export_market_closes(wb, conn):
    """Sheet 5 — Stored previous close prices."""
    ws = wb.create_sheet("Market Closes")

    cursor = conn.cursor()
    cursor.execute("""
        SELECT trade_date, ticker, close_price, created_at
        FROM market_closes
        ORDER BY trade_date DESC, ticker
    """)

    headers = ["Date", "Ticker", "Close Price", "Saved At"]
    ws.append(headers)
    style_header_row(ws, 1, color="7B3F00")

    rows = cursor.fetchall()
    for row in rows:
        ws.append(list(row))

    auto_width(ws)
    return len(rows)


def main():
    eastern = pytz.timezone(MARKET_TIMEZONE)
    now     = datetime.now(eastern)
    now_str = now.strftime("%Y-%m-%d %H:%M:%S %Z")
    date_str = now.strftime("%Y%m%d_%H%M")

    print(f"\n{'='*60}")
    print(f"  📊 DB EXPORT TO EXCEL")
    print(f"  {now_str}")
    print(f"{'='*60}\n")

    conn = get_db()
    wb   = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    print(f"  Exporting sheets...")

    count = export_paper_trades(wb, conn)
    print(f"  ✅ All Trades          — {count} rows")

    count = export_open_positions(wb, conn)
    print(f"  ✅ Open Positions      — {count} rows")

    count = export_closed_summary(wb, conn)
    print(f"  ✅ Closed Summary      — written")

    count = export_signals_today(wb, conn)
    print(f"  ✅ Today's Signals     — {count} rows")

    count = export_market_closes(wb, conn)
    print(f"  ✅ Market Closes       — {count} rows")

    conn.close()

    # Save file
    filename = f"scanner_export_{date_str}.xlsx"
    wb.save(filename)

    print(f"\n  💾 Saved: {filename}")
    print(f"\n{'='*60}\n")


if __name__ == "__main__":
    main()