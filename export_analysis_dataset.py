# export_analysis_dataset.py
"""
Exports a flat dataset for hurdle/trailing stop analysis in Excel.
One row per position snapshot, with full trade context on every row.
"""
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import pytz
from datetime import datetime
from dotenv import load_dotenv
load_dotenv()

def main():
    from journal import DB_PATH
    eastern = pytz.timezone("US/Eastern")
    now_str = datetime.now(eastern).strftime("%Y%m%d_%H%M")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    print("Loading data...")

    cursor.execute("""
        SELECT
            pt.id                       as trade_id,
            pt.signal_contract          as contract,
            pt.entry_date,
            pt.entry_time,
            pt.exit_date,
            pt.exit_time,
            pt.status,
            pt.exit_reason,
            pt.entry_price,
            pt.exit_price,
            pt.contracts,
            pt.total_cost,
            pt.pnl                      as final_pnl,
            pt.pnl_pct                  as final_pnl_pct,
            pt.hold_days,
            pt.score_at_entry,
            pt.dte_at_entry,
            pt.delta_at_entry,
            pt.iv_at_entry,
            pt.verdict_at_entry,
            pt.market_bias_at_entry,
            pt.target_price,
            pt.stop_price,
            pt.notes                    as trade_notes,
            ps.id                       as snapshot_id,
            ps.snapshot_time,
            ps.current_price,
            ps.bid,
            ps.ask,
            ps.pnl                      as snap_pnl,
            ps.pnl_pct                  as snap_pnl_pct,
            ps.dynamic_stop,
            ps.current_dte,
            ps.stop_triggered,
            ps.target_triggered,
            ps.spy_price,
            ps.spy_chg_pct,
            ps.qqq_price,
            ps.qqq_chg_pct,
            ps.vix_price
        FROM paper_trades pt
        LEFT JOIN position_snapshots ps ON ps.trade_id = pt.id
        ORDER BY pt.id ASC, ps.id ASC
    """)

    rows = cursor.fetchall()
    conn.close()

    print(f"Loaded {len(rows)} rows...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analysis Dataset"

    headers = [
        "trade_id", "contract", "entry_date", "entry_time",
        "exit_date", "exit_time", "status", "exit_reason",
        "entry_price", "exit_price", "contracts", "total_cost",
        "final_pnl", "final_pnl_pct", "hold_days",
        "score_at_entry", "dte_at_entry", "delta_at_entry",
        "iv_at_entry", "verdict_at_entry", "market_bias_at_entry",
        "target_price", "stop_price", "trade_notes",
        "snapshot_id", "snapshot_time", "current_price",
        "bid", "ask", "snap_pnl", "snap_pnl_pct",
        "dynamic_stop", "current_dte", "stop_triggered",
        "target_triggered", "spy_price", "spy_chg_pct",
        "qqq_price", "qqq_chg_pct", "vix_price"
    ]

    # Header row
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")

    # Data rows
    for row in rows:
        ws.append(list(row))

    # Auto-width
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value else 0 for cell in col),
            default=0
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

    ws.freeze_panes = "A2"

    filename = f"analysis_dataset_{now_str}.xlsx"
    wb.save(filename)
    print(f"Saved: {filename}")
    print(f"Rows: {len(rows)}")

if __name__ == "__main__":
    main()